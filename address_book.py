"""AddressBook Project"""
import os
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


class AddressBook:
    def __init__(self):
        pass

    def action(self, choice):
        choice = choice.lower()
        address_book_file = 'address_book.xlsx'

        if choice == 'load':
            if not os.path.isfile(address_book_file):
                # Create a new workbook and add columns
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(['name', 'phone number'])

                workbook.save(address_book_file)
                print(f'{address_book_file} created successfully!')
            else:
                print(f'{address_book_file} already exists.')

                # Load the existing workbook and display its content
                workbook = load_workbook(address_book_file)
                sheet = workbook.active

                for row in sheet.iter_rows(values_only=True):
                    print(row)
        elif choice == 'add':
            # Load the existing workbook
            workbook = load_workbook(address_book_file)
            sheet = workbook.active

            # Continuously prompt the user for new contacts until "end" is entered
            while True:
                name = input("Enter the name (or 'end' to finish): ")

                # Check if the user wants to finish entering contacts
                if name.lower() == 'end':
                    break

                phone_number = input("Enter the phone number: ")

                # Add the contact to the sheet
                sheet.append([name, phone_number])

                # Save the workbook
            workbook.save(address_book_file)
            print(f'{address_book_file} saved successfully!')
        elif choice == 'remove':
            workbook = load_workbook(address_book_file)
            sheet = workbook.active

            # Continuously prompt the user for contacts until "end" is entered
            while True:
                name = input("Enter the name (or 'end' to finish): ")

                # Check if the user wants to finish entering contacts
                if name.lower() == 'end':
                    break

                phone_number = input("Enter the phone number: ")

                # Check if the contact exists
                contact_exists = False
                for row in sheet.iter_rows(min_row=1, values_only=True):
                    if row[0] == name and row[1] == phone_number:
                        contact_exists = True
                        break

                if contact_exists:
                    # Remove the contact from the sheet
                    sheet.delete_rows(row[0].row)
                    print(f'{name} with phone number {phone_number} removed successfully.')
                else:
                    print(f'{name} with phone number {phone_number} does not exist in the address book.')

                # Save the workbook
            workbook.save(address_book_file)
            print(f'{address_book_file} saved successfully!')


address_book = AddressBook()
# print('Choose your action\n'
#       'Add Contact - Add\n'
#       'Remove Contact - Remove\n'
#       'Update Contact - Update\n'
#       'Search - Search\n'
#       'list all contact - List\n'
#       'Save - Save\n'
#       'Load - Load\n'
#       '\n')
choice = input('What do you want to do? -> ')
address_book.action(choice)